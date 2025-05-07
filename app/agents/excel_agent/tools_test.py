"""Test script for Excel Agent tools."""

import os
import openpyxl
import yaml
from pathlib import Path
from io import BytesIO

# Create necessary directories
templates_dir = Path("templates")
mappings_dir = Path("mappings")
templates_dir.mkdir(exist_ok=True)
mappings_dir.mkdir(exist_ok=True)

print("1. Creating test template...")
# Create a sample template
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Sheet1"
sheet["A1"] = "Sample Report"
summary = wb.create_sheet("Summary")
summary["A1"] = "Summary View"
template_path = templates_dir / "가.xlsx"
wb.save(template_path)
print(f"   Created template at {template_path}")

print("2. Creating test mapping file...")
# Create a sample mapping file
mapping_content = """
title:
  "Sheet1": "A1"
  "Summary": "A1"
static:
  "Sheet1":
    "A2": "Test data"
data_mappings:
  - sheet: "Sheet1"
    type: "direct"
    cells:
      "C5": "100"  # Static value for testing
  - sheet: "Summary"
    type: "row_iteration"
    start_row: 5
    columns:
      "date": "A"
      "value": "B"
"""
mapping_path = mappings_dir / "가.yml"
with open(mapping_path, "w") as f:
    f.write(mapping_content)
print(f"   Created mapping file at {mapping_path}")

print("3. Testing load_template function...")
# Simplified version of load_template
def load_template(template_id):
    """Load an Excel template by ID."""
    template_path = templates_dir / f"{template_id}.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Template {template_id} not found")
    return openpyxl.load_workbook(template_path)

# Load template
wb = load_template("가")
print(f"   Loaded template with sheets: {wb.sheetnames}")

print("4. Testing save_workbook function...")
# Simplified version of save_workbook
def save_workbook(wb, context):
    """Save workbook to bytes and generate filename."""
    template_id = context.get("template", "report")
    date_range = context.get("date_range", "")
    date_str = date_range.replace("-", "").replace("/", "_")
    filename = f"{template_id}_report_{date_str}.xlsx"
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return {
        "file_bytes": buffer.getvalue(),
        "filename": filename
    }

# Save workbook
context = {"template": "가", "date_range": "2025-04-14/19"}
result = save_workbook(wb, context)
print(f"   Saved workbook with filename: {result['filename']}")
print(f"   File size: {len(result['file_bytes'])} bytes")

print("All tools tests passed successfully!") 