"""Tools for the Excel Agent to manipulate workbooks and process data."""

import logging
import os
import yaml
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, Optional

import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from langchain.tools import tool

logger = logging.getLogger(__name__)

@tool
def load_template_wb(template_id: str) -> Workbook:
    """
    Load an Excel template by ID.
    
    Args:
        template_id: ID of the template to load (corresponds to filename)
        
    Returns:
        Loaded workbook object
        
    Raises:
        FileNotFoundError: If template does not exist
    """
    template_path = Path(f"app/agents/excel_agent/templates/{template_id}.xlsx")
    
    if not template_path.exists():
        logger.error(f"Template {template_id} not found at {template_path}")
        raise FileNotFoundError(f"Template {template_id} not found")
        
    logger.info(f"Loading template: {template_id}")
    return openpyxl.load_workbook(template_path)

@tool
def transform_df_for_template(
    df: pd.DataFrame, template_id: str, context: Dict[str, Any]
) -> pd.DataFrame:
    """
    Transform dataframe based on template requirements.
    
    Args:
        df: Source DataFrame
        template_id: ID of the template to prepare data for
        context: Additional context for transformation
        
    Returns:
        Transformed DataFrame ready for mapping
    """
    logger.info(f"Transforming dataframe for template: {template_id}")
    
    # Template-specific transformations
    if template_id == "가":
        # Example: Pivot data by date with sum aggregation
        if "date_column" in context:
            date_col = context["date_column"]
            if "pivot_column" in context and "value_column" in context:
                pivot_col = context["pivot_column"]
                value_col = context["value_column"]
                return pd.pivot_table(
                    df, 
                    index=date_col,
                    columns=pivot_col, 
                    values=value_col,
                    aggfunc="sum"
                ).reset_index()
    
    elif template_id == "나":
        # Example: Group by and aggregate
        if "group_columns" in context and "agg_columns" in context:
            group_cols = context["group_columns"]
            agg_dict = {col: "sum" for col in context["agg_columns"]}
            return df.groupby(group_cols).agg(agg_dict).reset_index()
    
    elif template_id == "다":
        # Example: Filter and sort
        if "filter_column" in context and "filter_value" in context:
            filtered = df[df[context["filter_column"]] == context["filter_value"]]
            if "sort_by" in context:
                return filtered.sort_values(context["sort_by"])
            return filtered
    
    elif template_id == "라":
        # Example: Time series resampling
        if "date_column" in context and "freq" in context:
            df[context["date_column"]] = pd.to_datetime(df[context["date_column"]])
            df = df.set_index(context["date_column"])
            if "value_column" in context:
                return df[context["value_column"]].resample(context["freq"]).sum().reset_index()
    
    elif template_id == "마":
        # Example: Calculate percentages
        if "total_column" in context and "part_columns" in context:
            total_col = context["total_column"]
            for col in context["part_columns"]:
                df[f"{col}_pct"] = (df[col] / df[total_col]) * 100
            return df
    
    # Default: return original DataFrame if no matching transformation
    logger.warning(f"No specific transformation for template {template_id}, returning original DataFrame")
    return df

@tool
def map_df_to_workbook(
    df: pd.DataFrame, wb: Workbook, template_id: str, context: Dict[str, Any]
) -> Workbook:
    """
    Map DataFrame values to workbook cells according to mapping rules.
    
    Args:
        df: Source DataFrame (transformed)
        wb: Target workbook
        template_id: ID of the template/mapping to use
        context: Additional context for mapping
        
    Returns:
        Updated workbook with data filled in
    """
    logger.info(f"Mapping data to workbook using template: {template_id}")
    
    # Load mapping rules from YAML file
    mapping_path = Path(f"app/agents/excel_agent/mappings/{template_id}.yml")
    
    if not mapping_path.exists():
        logger.error(f"Mapping file for template {template_id} not found")
        raise FileNotFoundError(f"Mapping file for template {template_id} not found")
    
    with open(mapping_path, "r") as f:
        mapping = yaml.safe_load(f)
    
    # Apply date range to title if specified
    if "date_range" in context and "title" in mapping:
        for sheet_name, title_cell in mapping["title"].items():
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                cell = sheet[title_cell]
                if "date_range" in context:
                    cell.value = f"{cell.value} {context['date_range']}"
    
    # Fill static cells (non-data-driven)
    if "static" in mapping:
        for sheet_name, cells in mapping["static"].items():
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for cell_addr, value in cells.items():
                    sheet[cell_addr] = value
    
    # Process data-driven mappings
    if "data_mappings" in mapping:
        for mapping_rule in mapping["data_mappings"]:
            sheet_name = mapping_rule.get("sheet")
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet {sheet_name} not found in workbook")
                continue
                
            sheet = wb[sheet_name]
            
            # Handle different mapping types
            mapping_type = mapping_rule.get("type", "direct")
            
            if mapping_type == "direct":
                _apply_direct_mapping(df, sheet, mapping_rule)
            elif mapping_type == "matrix":
                _apply_matrix_mapping(df, sheet, mapping_rule)
            elif mapping_type == "row_iteration":
                _apply_row_iteration_mapping(df, sheet, mapping_rule)
    
    return wb

def _apply_direct_mapping(df: pd.DataFrame, sheet: Worksheet, mapping_rule: Dict):
    """Apply direct cell-to-value mappings."""
    cell_mappings = mapping_rule.get("cells", {})
    for cell_addr, df_expr in cell_mappings.items():
        try:
            # Handle column reference
            if df_expr in df.columns:
                # Single value
                sheet[cell_addr] = df[df_expr].iloc[0]
            # Handle expressions with aggregate functions
            elif df_expr.startswith("sum(") and df_expr.endswith(")"):
                col_name = df_expr[4:-1]
                sheet[cell_addr] = df[col_name].sum()
            elif df_expr.startswith("mean(") and df_expr.endswith(")"):
                col_name = df_expr[5:-1]
                sheet[cell_addr] = df[col_name].mean()
            elif df_expr.startswith("max(") and df_expr.endswith(")"):
                col_name = df_expr[4:-1]
                sheet[cell_addr] = df[col_name].max()
            elif df_expr.startswith("min(") and df_expr.endswith(")"):
                col_name = df_expr[4:-1]
                sheet[cell_addr] = df[col_name].min()
            elif df_expr.startswith("count(") and df_expr.endswith(")"):
                col_name = df_expr[6:-1]
                sheet[cell_addr] = df[col_name].count()
        except Exception as e:
            logger.error(f"Error applying direct mapping to cell {cell_addr}: {str(e)}")

def _apply_matrix_mapping(df: pd.DataFrame, sheet: Worksheet, mapping_rule: Dict):
    """Apply matrix-style mappings (rows x columns)."""
    start_cell = mapping_rule.get("start_cell", "A1")
    row_col = mapping_rule.get("row_column")
    col_col = mapping_rule.get("column_column")
    value_col = mapping_rule.get("value_column")
    
    if not all([row_col, col_col, value_col]):
        logger.error("Matrix mapping requires row_column, column_column, and value_column")
        return
        
    # Create a pivot table for easier mapping
    try:
        pivot = df.pivot(index=row_col, columns=col_col, values=value_col)
        
        # Get starting row and column
        start_row = int("".join(filter(str.isdigit, start_cell)))
        start_col = openpyxl.utils.column_index_from_string(
            "".join(filter(str.isalpha, start_cell))
        )
        
        # Write column headers
        for i, col_name in enumerate(pivot.columns):
            sheet.cell(row=start_row, column=start_col + i + 1).value = col_name
            
        # Write row headers and values
        for i, (row_name, data) in enumerate(pivot.iterrows()):
            sheet.cell(row=start_row + i + 1, column=start_col).value = row_name
            for j, value in enumerate(data):
                if pd.notna(value):  # Skip NaN values
                    sheet.cell(row=start_row + i + 1, column=start_col + j + 1).value = value
    except Exception as e:
        logger.error(f"Error applying matrix mapping: {str(e)}")

def _apply_row_iteration_mapping(df: pd.DataFrame, sheet: Worksheet, mapping_rule: Dict):
    """Apply row iteration mappings (each DataFrame row to a row in the sheet)."""
    start_row = mapping_rule.get("start_row", 1)
    column_mapping = mapping_rule.get("columns", {})
    
    if not column_mapping:
        logger.error("Row iteration mapping requires 'columns' mapping")
        return
        
    # Convert column letters to indices
    col_mapping = {}
    for df_col, sheet_col in column_mapping.items():
        col_idx = openpyxl.utils.column_index_from_string(sheet_col)
        col_mapping[df_col] = col_idx
    
    # Write each row
    for i, row in enumerate(df.itertuples(index=False)):
        row_dict = row._asdict()
        curr_row = start_row + i
        
        for df_col, col_idx in col_mapping.items():
            if df_col in row_dict:
                sheet.cell(row=curr_row, column=col_idx).value = row_dict[df_col]

@tool
def save_workbook(wb: Workbook, context: Dict[str, Any]) -> Dict[str, Any]:
    """
    Save workbook to bytes and generate filename.
    
    Args:
        wb: Workbook to save
        context: Context with template and date range info
        
    Returns:
        Dictionary with file_bytes and filename
    """
    logger.info("Saving workbook to bytes")
    
    # Generate filename based on context
    template_id = context.get("template", "report")
    date_range = context.get("date_range", "")
    date_str = date_range.replace("-", "").replace("/", "_")
    filename = f"{template_id}_report_{date_str}.xlsx"
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return {
        "file_bytes": buffer.getvalue(),
        "filename": filename
    }

@tool
def upload_file(file_bytes: bytes, filename: str, upload_path: Optional[str] = None) -> str:
    """
    Upload Excel file to a specified location (optional).
    
    Args:
        file_bytes: Excel file as bytes
        filename: Filename to save as
        upload_path: Path to save file (optional)
        
    Returns:
        Path to saved file
    """
    if not upload_path:
        upload_path = os.environ.get("EXCEL_UPLOAD_PATH", "excel_exports")
        
    # Create directory if it doesn't exist
    os.makedirs(upload_path, exist_ok=True)
    
    # Save file
    file_path = f"{upload_path}/{filename}"
    with open(file_path, "wb") as f:
        f.write(file_bytes)
        
    logger.info(f"File saved to {file_path}")
    return file_path 