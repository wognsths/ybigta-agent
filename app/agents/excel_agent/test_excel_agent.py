"""Tests for the Excel Agent."""

import os
import pytest
import pandas as pd
import openpyxl
from io import BytesIO
from typing import Dict, Any

from app.agents.excel_agent.tools import (
    load_template_wb,
    transform_df_for_template,
    map_df_to_workbook,
    save_workbook,
)
from app.agents.excel_agent.graph import get_excel_agent, ExcelRequest, ExcelResponse

# Ensure template and mapping directories exist
os.makedirs("app/agents/excel_agent/templates", exist_ok=True)
os.makedirs("app/agents/excel_agent/mappings", exist_ok=True)

@pytest.fixture
def sample_df():
    """Create a sample DataFrame for testing."""
    return pd.DataFrame({
        "date": ["2025-04-14", "2025-04-15", "2025-04-16", "2025-04-17", "2025-04-18"],
        "category": ["A", "B", "A", "C", "B"],
        "value": [100, 150, 120, 200, 180],
        "quantity": [10, 15, 12, 20, 18],
    })

@pytest.fixture
def sample_context():
    """Create a sample context for testing."""
    return {
        "template": "가",
        "date_range": "2025-04-14/19",
        "date_column": "date",
        "pivot_column": "category",
        "value_column": "value",
    }

@pytest.fixture
def template_workbook():
    """Create a sample template workbook."""
    wb = openpyxl.Workbook()
    
    # Rename default sheet
    sheet = wb.active
    sheet.title = "Sheet1"
    
    # Add title
    sheet["A1"] = "Sample Report"
    
    # Add a second sheet
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Summary View"
    
    # Save template
    template_path = "app/agents/excel_agent/templates/가.xlsx"
    wb.save(template_path)
    
    return template_path

def test_load_template_wb(template_workbook):
    """Test loading a template workbook."""
    wb = load_template_wb("가")
    assert isinstance(wb, openpyxl.Workbook)
    assert "Sheet1" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    assert wb["Sheet1"]["A1"].value == "Sample Report"

def test_transform_df_for_template(sample_df, sample_context):
    """Test transforming a DataFrame for a template."""
    transformed_df = transform_df_for_template(sample_df, "가", sample_context)
    assert isinstance(transformed_df, pd.DataFrame)
    
    # Check if transformation is correct (pivot table)
    assert "date" in transformed_df.columns
    assert "A" in transformed_df.columns  # Pivoted category A
    assert "B" in transformed_df.columns  # Pivoted category B

def test_map_df_to_workbook(sample_df, sample_context, template_workbook):
    """Test mapping DataFrame to workbook."""
    # First create mapping file
    os.makedirs("app/agents/excel_agent/mappings", exist_ok=True)
    mapping_content = """
title:
  "Sheet1": "A1"
  "Summary": "A1"
static:
  "Sheet1":
    "A2": "Test data"
data_mappings:
  - sheet: "Sheet1"
    type: "direct"
    cells:
      "C5": "sum(value)"
  - sheet: "Summary"
    type: "row_iteration"
    start_row: 5
    columns:
      "date": "A"
      "value": "B"
"""
    with open("app/agents/excel_agent/mappings/가.yml", "w") as f:
        f.write(mapping_content)
    
    # Load and map
    wb = load_template_wb("가")
    updated_wb = map_df_to_workbook(sample_df, wb, "가", sample_context)
    
    # Check if mapping worked
    assert isinstance(updated_wb, openpyxl.Workbook)
    assert updated_wb["Sheet1"]["C5"].value == sample_df["value"].sum()

def test_save_workbook(sample_context, template_workbook):
    """Test saving a workbook to bytes."""
    wb = load_template_wb("가")
    result = save_workbook(wb, sample_context)
    
    assert "file_bytes" in result
    assert "filename" in result
    assert isinstance(result["file_bytes"], bytes)
    assert "가_report_" in result["filename"]

def test_excel_agent_workflow(sample_df, sample_context, template_workbook):
    """Test the entire Excel Agent workflow."""
    # Create the agent
    agent = get_excel_agent()
    
    # Execute the workflow
    result = agent.invoke({
        "df": sample_df,
        "context": sample_context
    })
    
    # Check result
    assert isinstance(result, ExcelResponse)
    assert hasattr(result, "file_bytes")
    assert hasattr(result, "filename")
    
    # Verify Excel file
    bytes_io = BytesIO(result.file_bytes)
    wb = openpyxl.load_workbook(bytes_io)
    
    assert "Sheet1" in wb.sheetnames
    assert "Summary" in wb.sheetnames 