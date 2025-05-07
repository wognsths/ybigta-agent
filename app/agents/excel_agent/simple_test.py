"""Simple test script for Excel Agent."""

import os
import pandas as pd
import openpyxl
from pathlib import Path

# Create directories if they don't exist
os.makedirs("templates", exist_ok=True)
os.makedirs("mappings", exist_ok=True)

# Create a simple DataFrame
df = pd.DataFrame({
    "date": ["2025-04-14", "2025-04-15", "2025-04-16"],
    "category": ["A", "B", "A"],
    "value": [100, 150, 120],
})

print("DataFrame created successfully:")
print(df)

# Create a simple Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Sheet1"
sheet["A1"] = "Test Report"

# Create a second sheet
summary = wb.create_sheet("Summary")
summary["A1"] = "Summary View"

# Save workbook
template_path = "templates/test.xlsx"
wb.save(template_path)

print(f"Workbook saved to {template_path}")

# Try to read the workbook back
wb2 = openpyxl.load_workbook(template_path)
print("Workbook loaded successfully")
print(f"Sheets: {wb2.sheetnames}")

print("All tests passed!") 