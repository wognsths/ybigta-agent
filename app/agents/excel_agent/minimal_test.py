"""Minimal test script for Excel Agent."""

import os
import openpyxl
from pathlib import Path

print("Starting minimal test...")

# Create directories if they don't exist
os.makedirs("templates", exist_ok=True)
print("Templates directory created.")

# Create a simple Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Sheet1"
sheet["A1"] = "Test Report"

# Create a second sheet
summary = wb.create_sheet("Summary")
summary["A1"] = "Summary View"

# Save workbook
template_path = "templates/test.xlsx"
wb.save(template_path)

print(f"Workbook saved to {template_path}")

# Try to read the workbook back
wb2 = openpyxl.load_workbook(template_path)
print("Workbook loaded successfully")
print(f"Sheets: {wb2.sheetnames}")

print("All tests passed!") 